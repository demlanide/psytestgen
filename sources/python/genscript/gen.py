import sqlite3
from openpyxl import load_workbook
import numpy as np
import math
from sklearn.linear_model import LinearRegression

limit = 4
update_limit_if_0 = 10
update_limit_percent = 0.9

connection = sqlite3.connect("/Users/dameli/Desktop/diploma/generator/sources/gendb.db")
crsr = connection.cursor()

sql_command = "SELECT count(*) FROM traits"
crsr.execute(sql_command)
ans = crsr.fetchall()
n = ans[0][0]

sql_command = "SELECT id_trait FROM traits"
crsr.execute(sql_command)
ans = crsr.fetchall()
lst = ans

wb = load_workbook("/Users/dameli/Desktop/diploma/generator/sources/analysis.xlsx")
title1 = 'regression_a'
title2 = 'regression_b'
title3 = 'correlation'

for i in lst:
    sql_command = "SELECT quantity FROM traits WHERE id_trait = " + str(i[0])
    crsr.execute(sql_command)
    ans = crsr.fetchall()
    quantity = int(ans[0][0])
    sql_command = "SELECT count(*) FROM results WHERE id_trait = " + str(i[0])
    crsr.execute(sql_command)
    ans = crsr.fetchall()
    total = int(ans[0][0])
    if (quantity == 0 and total > update_limit_if_0) or (total != 0 and quantity/total < update_limit_percent):
        wb[title1].cell(row=i[0], column=i[0]).value = 1
        wb[title2].cell(row=i[0], column=i[0]).value = 1
        wb[title3].cell(row=i[0], column=i[0]).value = 1
        for j in lst:
            if i == j:
                break
            else:
                sql_command = "SELECT a.result, b.result FROM results a INNER JOIN results b ON a.id_user = b.id_user WHERE a.id_trait = " + str(i[0]) + " AND b.id_trait = " + str(j[0]) + " AND a.result IS NOT NULL AND b.result IS NOT NULL AND a.result NOT LIKE '~%' AND b.result NOT LIKE '~%'"
                crsr.execute(sql_command)
                ans = crsr.fetchall()
                first_array = []
                second_array = []
                for k in range(0, len(ans)):
                    first_array.append(ans[k][0])
                    second_array.append(ans[k][1])
                if len(first_array) >= limit and len(second_array) >= limit:
                    y = np.corrcoef(first_array, second_array)
                    y = y[0][1]
                else:
                    y = 0
                if math.fabs(y) >= 0.7:
                    wb[title3].cell(row=i[0], column=j[0]).value = y
                    wb[title3].cell(row=j[0], column=i[0]).value = y
                    first_array = np.array(first_array).reshape((-1, 1))
                    second_array = np.array(second_array)
                    model = LinearRegression()
                    model.fit(first_array, second_array)
                    wb[title1].cell(row=i[0], column=j[0]).value = float(model.intercept_)
                    wb[title1].cell(row=j[0], column=i[0]).value = float(model.intercept_)
                    wb[title2].cell(row=i[0], column=j[0]).value = float(model.coef_)
                    wb[title2].cell(row=j[0], column=i[0]).value = float(model.coef_)
                else:
                    wb[title1].cell(row=i[0], column=j[0]).value = 'NULL'
                    wb[title1].cell(row=j[0], column=i[0]).value = 'NULL'
                    wb[title2].cell(row=i[0], column=j[0]).value = 'NULL'
                    wb[title2].cell(row=j[0], column=i[0]).value = 'NULL'
                    wb[title3].cell(row=i[0], column=j[0]).value = 0
                    wb[title3].cell(row=j[0], column=i[0]).value = 0

wb.save("/Users/dameli/Desktop/diploma/generator/sources/analysis.xlsx")