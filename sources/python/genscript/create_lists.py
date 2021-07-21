import sqlite3
from openpyxl import load_workbook
import sys


def find_all(x, lst, relations):
    for key in relations:
        if x in relations[key]:
            lst.append(key)
            find_all(key, lst, relations)


id_user = str(sys.argv[1])

connection = sqlite3.connect("/Users/dameli/Desktop/diploma/generator/sources/gendb.db")
crsr = connection.cursor()

sql_command = "SELECT count(*) FROM traits"
crsr.execute(sql_command)
ans = crsr.fetchall()
n = ans[0][0]

sql_command = "SELECT id_trait FROM traits"
crsr.execute(sql_command)
ans = crsr.fetchall()
lst = ans

wb = load_workbook("/Users/dameli/Desktop/diploma/generator/sources/analysis.xlsx")
title1 = 'regression_a'

relations = {}
for i in range(0, n):
    for j in range(i + 1, n):
        if wb[title1].cell(row=lst[i][0], column=lst[j][0]).value != 'NULL':
            try:
                relations[lst[i][0]].append(lst[j][0])
            except KeyError:
                relations[lst[i][0]] = [lst[j][0]]
            try:
                relations[lst[j][0]].append(lst[i][0])
            except KeyError:
                relations[lst[j][0]] = [lst[i][0]]
lst_new = []
crsr.execute("SELECT id_trait FROM results WHERE id_user = ? AND result IS NULL", id_user)
ans = crsr.fetchall()
for i in range(0, len(ans)):
    lst_new.append(int(ans[i][0]))
lst_new.sort()

lst = []
crsr.execute("SELECT id_trait FROM results WHERE id_user = ? AND result IS NOT NULL AND result NOT LIKE '~%'", id_user)
ans = crsr.fetchall()
for i in range(0, len(ans)):
    lst.append(int(ans[i][0]))

new_lst = {3: 1}

for key in new_lst.keys():
    yup = '~' + str(new_lst[key])
    sql_command = "UPDATE results SET result = '" + yup + "' WHERE id_user = " + str(id_user) + " AND id_trait = " + str(key)
    crsr.execute(sql_command)
    connection.commit()
