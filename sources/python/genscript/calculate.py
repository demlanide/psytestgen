import sqlite3
from openpyxl import load_workbook
import sys

id_user = str(sys.argv[1])

connection = sqlite3.connect("/Users/dameli/Desktop/diploma/generator/sources/gendb.db")
crsr = connection.cursor()

sql_command = "SELECT count(*) FROM traits"
crsr.execute(sql_command)
ans = crsr.fetchall()
n = ans[0][0]

sql_command = "SELECT id_trait FROM traits"
crsr.execute(sql_command)
ans = crsr.fetchall()
lst = ans

wb = load_workbook("/Users/dameli/Desktop/diploma/generator/sources/analysis.xlsx")
title1 = 'regression_a'  # названия листа в excel документе которая хранит а коэффициент
title2 = 'regression_b'  # название листа в excel документе которая хранит b коэффициент

sql_command = "SELECT id_trait, result FROM (SELECT id_user, id_trait, result FROM results WHERE result LIKE '~%') WHERE id_user = ?"
crsr.execute(sql_command, id_user)
ans = crsr.fetchall()
tests_num = []  # список номеров тестов которые нужно рассчитать
answers_evaluate = []   # список номеров тестов от результатов которых нужно рассчитать
for i in range(0, len(ans)):
    tests_num.append(ans[i][0])
    answers_evaluate.append(ans[i][1])

k = len(tests_num)

while k > 0:
    for i in range(0, len(tests_num)):
        x = int(answers_evaluate[i][1:])
        y = int(tests_num[i])
        a = wb[title1].cell(row=x, column=y).value
        b = wb[title2].cell(row=x, column=y).value
        sql_command = "SELECT result FROM results WHERE id_user = ? AND id_trait = " + str(x) + " AND result IS NOT NULL AND result NOT LIKE '~%'"
        crsr.execute(sql_command, id_user)
        ans = crsr.fetchall()
        if len(ans) != 0:
            test_res = ans[0][0]
            res = str(int(a + b * test_res))
            sql_command = "UPDATE results SET result = ? WHERE id_user = " + id_user + " AND id_trait = " + str(y)
            crsr.execute(sql_command, res)
            connection.commit()
            k -= 1
